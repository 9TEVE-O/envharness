# Copyright 2026 The EnvHarness Authors.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""SpreadsheetBenchEnv -- wrap SpreadsheetBench as an ActionableEnv.

Contract follows the EnvHarness ActionableEnv ABC: zero mutation logic here.
The only runtime this Bridge owns is a per-episode SANDBOX WORKING DIRECTORY
(a fresh temp folder holding a copy of the task's input spreadsheet). Every
agent action runs *inside that folder* -- the "all operations confined to one
folder" model, the file-based analogue of the swebench Docker container or the
alfworld TextWorld engine. Rules-written code never sees the working dir, only
the `SpreadsheetBenchEnvState` dataclass (pure data).

Action contract (multi-turn ReAct, two tools):
    run_python(code: str) -- execute a Python snippet in the working dir
                             (openpyxl/pandas available); the agent must save
                             its result to `output_path`. Stateless per call.
    submit()              -- declare output_path final; ends the episode and
                             triggers grading.
    If the agent never submits, the episode runs to max_steps and we grade
    whatever is at output_path (it is pre-seeded as a copy of the input).

Evaluation (Online-Judge, single test case for verified_400):
    `evaluate()` recalculates the agent's output_path with LibreOffice (so
    formulas have cached values), recalculates the golden once (cached), then
    runs the official cell-value comparison over `answer_position`. success =
    every answer cell matches.

Setup requirements (one-time on the host):
    pip install openpyxl pandas
    sudo apt install libreoffice-calc        # headless formula recalculation
    # data: a `spreadsheetbench_verified_400/` dir (dataset.json + spreadsheet/)

Reset options (passed through `options` in reset()):
    data_path: str       -- dir with dataset.json + spreadsheet/. Falls back to
                            $SPREADSHEETBENCH_DATA.
    instance_id: str     -- explicit SpreadsheetBench id (e.g. "13-1"). If set,
                            takes priority over seed-based selection. Used by the
                            eval driver, which drives the env directly.
    sandbox_root: str    -- parent dir for per-episode working dirs (default a
                            system temp dir).
    soffice_path: str    -- explicit LibreOffice binary.
    python_exe: str      -- interpreter for run_python (default sys.executable).
    step_timeout: int    -- per-run_python timeout, seconds (default 60).
    obs_truncate: int    -- max observation chars (default 6000).
    preview_rows/cols: int -- size of the spreadsheet preview in the obs.
    recalc_golden: bool  -- recalc the golden before comparison (default True;
                            cached per golden path).
"""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from dataclasses import dataclass, field
from typing import Any, ClassVar

from envharness.core.actionable_env import ActionableEnv
from envharness.core.registry import register_env
from envharness.core.tool import Tool
from envharness.core.types import (
    Action, EnvResetResponse, EnvResponse, EvaluationResult, Observation,
    TaskSummary,
)
from . import online_judge_eval
from .dataset import (
    DATA_PATH_ENV, SBTask, load_dataset, select_task,
)
from .tools import RunPython, Submit


DEFAULT_STEP_TIMEOUT = 60
DEFAULT_OBS_TRUNCATE = 6000
DEFAULT_PREVIEW_ROWS = 20
DEFAULT_PREVIEW_COLS = 16


@dataclass
class SpreadsheetBenchEnvState:
    """Bridge-exposed view for Rules hooks. Pure data -- no file handles, no
    subprocess objects. Rules may read every field; the only writable surface
    is `extras` (per-episode scratch for Mutator-emitted hook state)."""
    task_id: str = ""                       # SpreadsheetBench id (e.g. "13-1")
    instruction: str = ""                   # the manipulation request
    instruction_type: str = ""              # Cell-Level | Sheet-Level Manipulation
    answer_position: str = ""               # graded cells (e.g. "A3:D32")
    answer_sheet: str = ""                  # sheet name for Sheet-Level tasks
    input_path: str = ""                    # abs path of the working-dir input copy
    output_path: str = ""                   # abs path the agent must write to
    spreadsheet_preview: str = ""           # first rows×cols rendered as text
    last_code: str = ""                     # most recent run_python snippet
    last_output: str = ""                   # most recent stdout+stderr (truncated)
    last_returncode: int = 0                # most recent run_python returncode
    submitted: bool = False                 # True once the agent called submit()
    step_count: int = 0
    extras: dict[str, Any] = field(default_factory=dict)


@register_env("spreadsheetbench")
class SpreadsheetBenchEnv(ActionableEnv):
    """SpreadsheetBench ActionableEnv. Per-episode sandbox working dir;
    multi-turn run_python ReAct; LibreOffice Online-Judge grading."""

    tool_registry: ClassVar[list[type[Tool]]] = [RunPython, Submit]

    def __init__(self) -> None:
        super().__init__()
        self._data_path: str = ""
        self._task: SBTask | None = None
        self._workdir: str | None = None
        self._sandbox_root: str | None = None
        self._soffice_path: str | None = None
        self._python_exe: str = sys.executable
        self._step_timeout: int = DEFAULT_STEP_TIMEOUT
        self._obs_truncate: int = DEFAULT_OBS_TRUNCATE
        self._preview_rows: int = DEFAULT_PREVIEW_ROWS
        self._preview_cols: int = DEFAULT_PREVIEW_COLS
        self._recalc_golden: bool = True
        self._terminated: bool = False
        self._eval_cache: EvaluationResult | None = None
        self.state: SpreadsheetBenchEnvState = SpreadsheetBenchEnvState()
        # Retained for save/load round-trips.
        self._last_reset_seed: int | None = None
        self._last_reset_options: dict[str, Any] = {}

    # -- core env interface -------------------------------------------------

    def reset(self, seed: int | None = None,
              options: dict[str, Any] | None = None) -> EnvResetResponse:
        opts = options or {}
        self._last_reset_seed = seed
        self._last_reset_options = dict(opts)

        self._data_path = (opts.get("data_path")
                           or os.environ.get(DATA_PATH_ENV) or "")
        if not self._data_path:
            raise RuntimeError(
                "SpreadsheetBenchEnv.reset: no data_path. Set "
                f"reset_options['data_path'] or ${DATA_PATH_ENV}."
            )
        self._sandbox_root = opts.get("sandbox_root") or None
        self._soffice_path = opts.get("soffice_path")
        self._python_exe = opts.get("python_exe") or sys.executable
        self._step_timeout = int(opts.get("step_timeout") or DEFAULT_STEP_TIMEOUT)
        self._obs_truncate = int(opts.get("obs_truncate") or DEFAULT_OBS_TRUNCATE)
        self._preview_rows = int(opts.get("preview_rows") or DEFAULT_PREVIEW_ROWS)
        self._preview_cols = int(opts.get("preview_cols") or DEFAULT_PREVIEW_COLS)
        self._recalc_golden = bool(opts.get("recalc_golden", True))

        # `instance_id` (explicit id, used by the eval driver) wins; otherwise
        # select by seed. The orchestrator-injected options["task_id"] is the
        # project label, NOT an instance -- see data.select_task.
        instance_id = opts.get("instance_id") or None
        task, _idx = select_task(self._data_path, seed, instance_id,
                                 multi=bool(opts.get("multi_test_case")))
        self._task = task

        if not task.init_path or not os.path.isfile(task.init_path):
            raise RuntimeError(
                f"SpreadsheetBench task {task.id!r}: input spreadsheet not "
                f"found (init_path={task.init_path!r})."
            )

        # Fresh sandbox working dir; copy the input in, pre-seed the output as a
        # copy so a no-op agent still leaves a gradeable file at output_path.
        self._teardown_workdir()
        self._workdir = tempfile.mkdtemp(prefix=f"sb_{task.id}_",
                                         dir=self._sandbox_root)
        input_path = os.path.join(self._workdir, f"{task.id}_input.xlsx")
        output_path = os.path.join(self._workdir, f"{task.id}_output.xlsx")
        shutil.copyfile(task.init_path, input_path)
        shutil.copyfile(task.init_path, output_path)

        self._terminated = False
        self._eval_cache = None
        self.state = SpreadsheetBenchEnvState(
            task_id=task.id,
            instruction=task.instruction,
            instruction_type=task.instruction_type,
            answer_position=task.answer_position,
            answer_sheet=task.answer_sheet,
            input_path=input_path,
            output_path=output_path,
            spreadsheet_preview=self._build_preview(input_path),
            step_count=0,
        )
        return EnvResetResponse(
            observation=self._observe(),
            info={
                "task_id": task.id,
                "instruction_type": task.instruction_type,
                "answer_position": task.answer_position,
                "won": False,
            },
        )

    def step(self, action: Action) -> EnvResponse:
        if self._terminated:
            # Defensive: the runner stops on terminated, but a stray post-term
            # step should be a harmless no-op rather than re-running code.
            return EnvResponse(
                observation=self._observe(), reward=0.0,
                terminated=True, truncated=False, info={"won": False},
            )

        if action.name == Submit.name:
            self.state.submitted = True
            self._terminated = True
            return EnvResponse(
                observation=Observation(
                    text="[submitted] output_path will be graded.",
                    data={"submitted": True},
                ),
                reward=0.0, terminated=True, truncated=False,
                info={"submitted": True, "won": None},
            )

        if action.name != RunPython.name:
            return EnvResponse(
                observation=Observation(
                    text=f"[unknown tool: {action.name}] valid tools: "
                         "run_python(code), submit()",
                    data={"error": "unknown_tool"},
                ),
                reward=0.0, terminated=False, truncated=False,
                info={"error": "unknown_tool"},
            )

        code = action.kwargs.get("code")
        if code is None:
            code = action.kwargs.get("text") or ""
        code = str(code)
        if not code.strip():
            return EnvResponse(
                observation=Observation(
                    text="[empty code] run_python needs a non-empty `code` arg.",
                    data={"error": "empty_code"},
                ),
                reward=0.0, terminated=False, truncated=False,
                info={"error": "empty_code"},
            )

        self.state.step_count += 1
        stdout, returncode = self._run_python(code)
        self.state.last_code = code
        self.state.last_output = stdout
        self.state.last_returncode = returncode

        return EnvResponse(
            observation=self._observe(run_output=stdout),
            reward=0.0, terminated=False, truncated=False,
            info={
                "result": {"returncode": returncode,
                           "output": stdout[:1000]},
                "returncode": returncode,
                "won": None,
            },
        )

    def evaluate(self) -> EvaluationResult:
        if self._eval_cache is not None:
            return self._eval_cache
        task = self._task
        if task is None or not self._workdir:
            return EvaluationResult(success=False, score=0.0,
                                    metrics={"error": "no active episode"})
        result = self._grade(task)
        self._eval_cache = result
        return result

    def observe(self) -> Observation:
        return self._observe()

    def get_env_state(self) -> SpreadsheetBenchEnvState:
        return self.state

    @classmethod
    def env_state_schema(cls) -> str:
        return (
            "env_state is a SpreadsheetBenchEnvState dataclass with these fields:\n"
            "  task_id: str               -- SpreadsheetBench instance id\n"
            "  instruction: str           -- the spreadsheet manipulation request\n"
            "  instruction_type: str      -- 'Cell-Level Manipulation' | 'Sheet-Level Manipulation'\n"
            "  answer_position: str       -- graded cell range, e.g. 'A3:D32' or \"Sheet!B2\"\n"
            "  answer_sheet: str          -- sheet name for Sheet-Level tasks ('' otherwise)\n"
            "  input_path: str            -- abs path of the input spreadsheet copy\n"
            "  output_path: str           -- abs path the agent must write its answer to\n"
            "  spreadsheet_preview: str   -- first rows x cols of the input rendered as text\n"
            "  last_code: str             -- most recent run_python snippet\n"
            "  last_output: str           -- most recent stdout+stderr (truncated)\n"
            "  last_returncode: int       -- most recent run_python exit code\n"
            "  submitted: bool            -- True once the agent called submit()\n"
            "  step_count: int            -- run_python actions taken so far\n"
            "  extras: dict               -- free scratch dict for Rules hook state\n"
            "S0 changes go through in_env_actions (Setup layer), not Rules. Rules hooks may\n"
            "read all fields and write to `extras`. To change what the agent sees, return a\n"
            "new Observation from filter_observation -- do NOT mutate env_state fields."
        )

    # -- save / load --------------------------------------------------------
    #
    # Like alfworld/swebench: the working dir cannot be cheaply snapshotted, so
    # save/load captures only (seed, options) to reconstruct AT EPISODE
    # BOUNDARIES. Loading returns a fresh instance; the caller calls reset().
    # Mid-episode action history is the Setup harness's responsibility.

    def save_state(self) -> dict:
        return {
            "reset_seed": self._last_reset_seed,
            "reset_options": dict(self._last_reset_options),
        }

    @classmethod
    def from_state(cls, state: dict) -> "SpreadsheetBenchEnv":
        env = cls()
        env._last_reset_seed = state.get("reset_seed")
        env._last_reset_options = dict(state.get("reset_options") or {})
        return env

    def default_reset_args(self) -> tuple[int | None, dict]:
        return self._last_reset_seed, dict(self._last_reset_options)

    # -- agent-driven task selection --------------------------------------

    @classmethod
    def list_tasks(cls, reset_options: dict | None = None,
                   limit: int | None = None) -> list[TaskSummary]:
        opts = dict(reset_options or {})
        data_path = opts.get("data_path") or os.environ.get(DATA_PATH_ENV) or ""
        if not data_path:
            raise RuntimeError(
                "SpreadsheetBenchEnv.list_tasks: no data_path "
                f"(reset_options['data_path'] or ${DATA_PATH_ENV})."
            )
        tasks = load_dataset(data_path)
        n = limit if limit is not None else len(tasks)
        out: list[TaskSummary] = []
        for i, t in enumerate(tasks[:n]):
            out.append(TaskSummary(
                task_idx=i,
                instance_id=t.id,
                brief=f"[{t.instruction_type}] {t.instruction[:180]}",
                metadata={"answer_position": t.answer_position,
                          "excluded": t.excluded,
                          "missing": t.missing},
            ))
        return out

    # -- close --------------------------------------------------------------

    def close(self) -> None:
        super().close()
        self._teardown_workdir()

    # -- helpers ------------------------------------------------------------

    def _teardown_workdir(self) -> None:
        if self._workdir and os.path.isdir(self._workdir):
            shutil.rmtree(self._workdir, ignore_errors=True)
        self._workdir = None

    def _run_python(self, code: str) -> tuple[str, int]:
        """Write `code` to a unique file in the working dir and execute it with
        `python_exe` (cwd = working dir). Return (combined_output, returncode)."""
        assert self._workdir is not None
        # Write the snippet to a scratch file OUTSIDE the working dir (cwd stays
        # the working dir, so the agent's relative paths still resolve). This
        # keeps the agent's own `os.listdir(working_dir)` clean -- otherwise it
        # sees a pile of `_step_*.py` files and wastes turns inspecting them /
        # hunting for "grading logic".
        fd, script = tempfile.mkstemp(suffix=".py", prefix="sbstep_")
        try:
            with os.fdopen(fd, "w") as f:
                f.write(code)
            proc = subprocess.run(
                [self._python_exe, script],
                cwd=self._workdir, capture_output=True, text=True,
                timeout=self._step_timeout,
            )
        except subprocess.TimeoutExpired:
            return (f"[ERROR] run_python timed out after {self._step_timeout}s",
                    -1)
        except Exception as e:  # noqa: BLE001
            return (f"[ERROR] failed to execute: {type(e).__name__}: {e}", -1)
        finally:
            try:
                os.remove(script)
            except OSError:
                pass
        parts = []
        if proc.stdout:
            parts.append(proc.stdout)
        if proc.stderr:
            parts.append(f"[STDERR]\n{proc.stderr}")
        if proc.returncode != 0:
            parts.append(f"[exit code: {proc.returncode}]")
        out = "\n".join(parts).strip() or "[run_python completed with no output]"
        if len(out) > self._obs_truncate:
            out = out[: self._obs_truncate] + "\n...[output truncated]"
        return out, proc.returncode

    def _grade(self, task: SBTask) -> EvaluationResult:
        """Recalc the agent output (and golden), run the official cell compare.
        success = all answer-position cells match the golden."""
        assert self._workdir is not None
        output_path = self.state.output_path
        if not os.path.isfile(output_path):
            return EvaluationResult(success=False, score=0.0,
                                    metrics={"error": "output not written"})
        # Recalc the agent's output in place (it is disposable). A recalc
        # FAILURE (timeout / soffice nonzero / missing convert) must NOT be
        # scored as a policy failure: under data_only=True the formula cells
        # would read None and a CORRECT episode would be judged fail. Raise so
        # it surfaces as rec["error"] (a visible eval_error, excluded from SR by
        # reasoning_bank_eval's summary + re-run by the dispatcher sweep; note reasoning_bank_eval.py's
        # own _run_condition does NOT resume) instead of a silent policy fail.
        if not online_judge_eval.recalc_with_libreoffice(output_path, self._soffice_path):
            raise RuntimeError(
                "eval_error: LibreOffice recalc of agent output failed "
                "(transient); not a policy failure")
        # Recalc the golden once, into a cached copy (golden is shared/read-only).
        golden = task.golden_path
        if self._recalc_golden:
            golden = self._cached_golden(task)
        passed, msg = online_judge_eval.compare_workbooks(
            golden, output_path, task.instruction_type, task.answer_position,
        )
        return EvaluationResult(
            success=bool(passed),
            score=1.0 if passed else 0.0,
            metrics={"won": bool(passed), "diff": msg[:300],
                     "answer_position": task.answer_position,
                     "submitted": self.state.submitted,
                     "steps": self.state.step_count},
        )

    def _cached_golden(self, task: SBTask) -> str:
        """Return a recalc'd copy of the golden, cached under the sandbox root
        keyed by task id so K rollouts / repeated evals pay LibreOffice once."""
        cache_dir = os.path.join(self._sandbox_root or tempfile.gettempdir(),
                                 "_sb_golden_cache")
        os.makedirs(cache_dir, exist_ok=True)
        cached = os.path.join(cache_dir, f"{task.id}_golden.xlsx")
        if not os.path.isfile(cached):
            tmp = cached + f".{uuid.uuid4().hex[:6]}.tmp"
            shutil.copyfile(task.golden_path, tmp)
            try:
                # recalc returns False (NO exception) on transient failure; the
                # old `except Exception` guard never fired, so os.replace cached
                # an un-recalc'd golden -> that task mis-scored in EVERY future
                # run (formula cells read None). Refuse to cache on failure and
                # raise so the episode surfaces as a visible eval_error (excluded
                # from SR; re-run only under the dispatcher sweep, not reasoning_bank_eval.py).
                if not online_judge_eval.recalc_with_libreoffice(tmp, self._soffice_path):
                    raise RuntimeError(
                        "eval_error: LibreOffice recalc of golden failed "
                        "(transient); refusing to cache a bad golden")
                os.replace(tmp, cached)
            finally:
                if os.path.isfile(tmp):
                    os.remove(tmp)
        return cached

    def _build_preview(self, xlsx_path: str) -> str:
        """Render the first preview_rows x preview_cols of each sheet as text.
        Best-effort: tries cached values (data_only) and falls back to the raw
        (formula) view. Never raises -- a preview failure must not abort reset."""
        try:
            import openpyxl
        except ImportError:
            return "(openpyxl not installed; no preview)"
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            return f"(could not open spreadsheet for preview: {e})"
        lines: list[str] = []
        try:
            for ws in wb.worksheets:
                lines.append(f"### sheet: {ws.title}  "
                             f"(dims {ws.max_row} rows x {ws.max_column} cols)")
                r = 0
                for row in ws.iter_rows(max_row=self._preview_rows,
                                        max_col=self._preview_cols,
                                        values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    lines.append("\t".join(cells))
                    r += 1
                    if r >= self._preview_rows:
                        break
                lines.append("")
        finally:
            wb.close()
        return "\n".join(lines).strip()

    def _observe(self, run_output: str | None = None) -> Observation:
        s = self.state
        parts = [
            f"working_directory: {self._workdir}",
            f"instruction: {s.instruction}",
            f"spreadsheet_path (input): {s.input_path}",
            f"output_path (write your answer here): {s.output_path}",
            f"instruction_type: {s.instruction_type}",
            f"answer_position: {s.answer_position}",
        ]
        if s.answer_sheet:
            parts.append(f"answer_sheet: {s.answer_sheet}")
        parts.append("spreadsheet_content (preview of the input):\n"
                     + s.spreadsheet_preview)
        if run_output is not None:
            parts.append("last run_python output:\n" + run_output)
        text = "\n\n".join(parts)
        return Observation(
            text=text,
            data={
                "task_id": s.task_id,
                "instruction": s.instruction,
                "instruction_type": s.instruction_type,
                "answer_position": s.answer_position,
                "answer_sheet": s.answer_sheet,
                "input_path": s.input_path,
                "output_path": s.output_path,
                "working_directory": self._workdir,
                "step_count": s.step_count,
                "last_returncode": s.last_returncode,
                "submitted": s.submitted,
            },
        )
